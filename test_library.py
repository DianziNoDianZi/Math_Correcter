"""
试卷库管理模块 - 增强版
支持考试管理、答题卡扫描、成绩审核等完整教学流程
"""
import os
import uuid
import json
import logging
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
from collections import defaultdict

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 目录配置
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
TEST_LIBRARY_DIR = os.path.join(DATA_DIR, 'test_library')
CLASSES_DIR = os.path.join(DATA_DIR, 'classes')
EXAMS_DIR = os.path.join(DATA_DIR, 'exams')
REPORTS_DIR = os.path.join(DATA_DIR, 'reports')

# 确保目录存在
for d in [DATA_DIR, TEST_LIBRARY_DIR, CLASSES_DIR, EXAMS_DIR, REPORTS_DIR]:
    os.makedirs(d, exist_ok=True)
    os.makedirs(os.path.join(d, 'images'), exist_ok=True)
    os.makedirs(os.path.join(d, 'scans'), exist_ok=True)

# ========== 配置文件读写 ==========

def load_json(filepath: str) -> Dict:
    """加载JSON文件"""
    try:
        if os.path.exists(filepath):
            with open(filepath, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {}
    except Exception as e:
        logger.error(f'加载JSON失败: {e}')
        return {}

def save_json(filepath: str, data: Dict) -> bool:
    """保存JSON文件"""
    try:
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        logger.error(f'保存JSON失败: {e}')
        return False

# ========== 试卷库元数据 ==========

def load_library_metadata() -> Dict:
    """加载试卷库元数据"""
    return load_json(os.path.join(TEST_LIBRARY_DIR, 'metadata.json'))

def save_library_metadata(metadata: Dict) -> bool:
    """保存试卷库元数据"""
    return save_json(os.path.join(TEST_LIBRARY_DIR, 'metadata.json'), metadata)

def init_library_metadata() -> Dict:
    """初始化试卷库元数据"""
    metadata = load_library_metadata()
    if not metadata:
        metadata = {
            'papers': [],
            'total_uploaded': 0,
            'last_updated': datetime.now().isoformat()
        }
        save_library_metadata(metadata)
    return metadata

# ========== 班级元数据 ==========

def load_classes_metadata() -> Dict:
    """加载班级元数据"""
    return load_json(os.path.join(CLASSES_DIR, 'metadata.json'))

def save_classes_metadata(metadata: Dict) -> bool:
    """保存班级元数据"""
    return save_json(os.path.join(CLASSES_DIR, 'metadata.json'), metadata)

def init_classes_metadata() -> Dict:
    """初始化班级元数据"""
    metadata = load_classes_metadata()
    if not metadata:
        metadata = {
            'classes': [],
            'total_students': 0
        }
        save_classes_metadata(metadata)
    return metadata

# ========== 考试元数据 ==========

def load_exams_metadata() -> Dict:
    """加载考试元数据"""
    return load_json(os.path.join(EXAMS_DIR, 'metadata.json'))

def save_exams_metadata(metadata: Dict) -> bool:
    """保存考试元数据"""
    return save_json(os.path.join(EXAMS_DIR, 'metadata.json'), metadata)

def init_exams_metadata() -> Dict:
    """初始化考试元数据"""
    metadata = load_exams_metadata()
    if not metadata:
        metadata = {'exams': []}
        save_exams_metadata(metadata)
    return metadata

# ========== 试卷库管理 ==========

def get_all_papers() -> List[Dict[str, Any]]:
    """获取所有试卷"""
    metadata = load_library_metadata()
    return metadata.get('papers', [])

def get_paper_by_id(paper_id: str) -> Optional[Dict[str, Any]]:
    """根据ID获取试卷"""
    metadata = load_library_metadata()
    for paper in metadata.get('papers', []):
        if paper['id'] == paper_id:
            return paper
    return None

def load_paper_analysis(paper_id: str) -> Optional[Dict[str, Any]]:
    """加载试卷分析结果"""
    analysis_path = os.path.join(TEST_LIBRARY_DIR, 'analyses', f'{paper_id}.json')
    return load_json(analysis_path)

def save_paper_analysis(paper_id: str, analysis: Dict) -> bool:
    """保存试卷分析结果"""
    analysis_dir = os.path.join(TEST_LIBRARY_DIR, 'analyses')
    os.makedirs(analysis_dir, exist_ok=True)
    analysis_path = os.path.join(analysis_dir, f'{paper_id}.json')
    return save_json(analysis_path, analysis)

# ========== 班级管理 ==========

def get_all_classes() -> List[Dict[str, Any]]:
    """获取所有班级"""
    metadata = load_classes_metadata()
    return metadata.get('classes', [])

def get_class_by_id(class_id: str) -> Optional[Dict[str, Any]]:
    """根据ID获取班级"""
    metadata = load_classes_metadata()
    for cls in metadata.get('classes', []):
        if cls['id'] == class_id:
            return cls
    return None

def add_class(class_data: Dict[str, Any]) -> Dict[str, Any]:
    """添加班级"""
    class_id = str(uuid.uuid4())
    
    metadata = load_classes_metadata()
    new_class = {
        'id': class_id,
        'name': class_data.get('name', '未命名班级'),
        'grade': class_data.get('grade', '10-12'),
        'teacher_name': class_data.get('teacher_name', ''),
        'created_at': datetime.now().isoformat(),
        'students': [],
        'assigned_papers': []
    }
    
    metadata['classes'].append(new_class)
    save_classes_metadata(metadata)
    
    return {
        'success': True,
        'class_id': class_id,
        'class': new_class
    }

def add_student_to_class(class_id: str, student_data: Dict[str, Any]) -> bool:
    """添加学生到班级"""
    metadata = load_classes_metadata()
    for cls in metadata['classes']:
        if cls['id'] == class_id:
            student_id = str(uuid.uuid4())
            student = {
                'id': student_id,
                'name': student_data.get('name', ''),
                'student_number': student_data.get('student_number', ''),
                'added_at': datetime.now().isoformat(),
                'scores': []
            }
            cls['students'].append(student)
            metadata['total_students'] = metadata.get('total_students', 0) + 1
            save_classes_metadata(metadata)
            return True
    return False

def assign_paper_to_class(class_id: str, paper_id: str) -> bool:
    """将试卷关联到班级"""
    metadata = load_classes_metadata()
    for cls in metadata['classes']:
        if cls['id'] == class_id:
            if paper_id not in cls['assigned_papers']:
                cls['assigned_papers'].append(paper_id)
                save_classes_metadata(metadata)
            return True
    return False

def delete_class(class_id: str) -> bool:
    """删除班级"""
    metadata = load_classes_metadata()
    initial_length = len(metadata['classes'])
    metadata['classes'] = [c for c in metadata['classes'] if c['id'] != class_id]
    
    if len(metadata['classes']) < initial_length:
        save_classes_metadata(metadata)
        return True
    return False

# ========== 学生成绩管理 ==========

def add_student_score(class_id: str, student_id: str, score_data: Dict[str, Any]) -> bool:
    """添加或更新学生成绩"""
    metadata = load_classes_metadata()
    for cls in metadata['classes']:
        if cls['id'] == class_id:
            for student in cls['students']:
                if student['id'] == student_id:
                    if 'scores' not in student:
                        student['scores'] = []
                    
                    score_record = {
                        'paper_id': score_data.get('paper_id'),
                        'paper_name': score_data.get('paper_name', '未命名试卷'),
                        'score': float(score_data.get('score', 0)),
                        'total_score': float(score_data.get('total_score', 100)),
                        'recorded_at': datetime.now().isoformat(),
                        'notes': score_data.get('notes', '')
                    }
                    student['scores'].append(score_record)
                    student['last_score'] = score_record
                    save_classes_metadata(metadata)
                    return True
    return False

def get_student_scores(class_id: str, student_id: str) -> List[Dict[str, Any]]:
    """获取学生的所有成绩"""
    metadata = load_classes_metadata()
    for cls in metadata['classes']:
        if cls['id'] == class_id:
            for student in cls['students']:
                if student['id'] == student_id:
                    return student.get('scores', [])
    return []

def get_class_scores(class_id: str, paper_id: str = None) -> List[Dict[str, Any]]:
    """获取班级的所有成绩"""
    cls = get_class_by_id(class_id)
    if not cls:
        return []
    
    all_scores = []
    for student in cls.get('students', []):
        for score in student.get('scores', []):
            if paper_id is None or score.get('paper_id') == paper_id:
                all_scores.append({
                    'student_id': student['id'],
                    'student_name': student.get('name', '未知'),
                    'student_number': student.get('student_number', ''),
                    **score
                })
    return all_scores

def calculate_class_statistics(class_id: str, paper_id: str = None) -> Dict[str, Any]:
    """计算班级成绩统计"""
    scores = get_class_scores(class_id, paper_id)
    
    if not scores:
        return {'total_students': 0, 'participated': 0, 'statistics': {}}
    
    score_values = [s['score'] for s in scores]
    percentages = [s['score'] / s['total_score'] * 100 if s['total_score'] > 0 else 0 for s in scores]
    
    statistics = {
        'total_students': len(get_class_by_id(class_id).get('students', [])),
        'participated': len(scores),
        'average_score': sum(score_values) / len(score_values) if score_values else 0,
        'average_percentage': sum(percentages) / len(percentages) if percentages else 0,
        'highest_score': max(score_values) if score_values else 0,
        'lowest_score': min(score_values) if score_values else 0,
        'pass_count': sum(1 for p in percentages if p >= 60),
        'pass_rate': sum(1 for p in percentages if p >= 60) / len(percentages) * 100 if percentages else 0,
        'excellent_count': sum(1 for p in percentages if p >= 90),
        'excellent_rate': sum(1 for p in percentages if p >= 90) / len(percentages) * 100 if percentages else 0,
    }
    
    return {
        'total_students': statistics['total_students'],
        'participated': statistics['participated'],
        'statistics': statistics
    }

def get_student_progress(class_id: str, student_id: str) -> Dict[str, Any]:
    """获取学生进步情况"""
    scores = get_student_scores(class_id, student_id)
    
    if len(scores) < 2:
        return {'has_progress': False, 'message': '成绩记录不足，无法分析进步情况'}
    
    sorted_scores = sorted(scores, key=lambda x: x.get('recorded_at', ''))
    
    recent_count = min(3, len(sorted_scores))
    recent_avg = sum(s['score'] / s['total_score'] * 100 for s in sorted_scores[-recent_count:]) / recent_count
    
    first_count = min(2, len(sorted_scores))
    first_avg = sum(s['score'] / s['total_score'] * 100 for s in sorted_scores[:first_count]) / first_count
    
    progress = recent_avg - first_avg
    
    return {
        'has_progress': True,
        'student_id': student_id,
        'total_exams': len(scores),
        'first_average': round(first_avg, 1),
        'recent_average': round(recent_avg, 1),
        'progress': round(progress, 1),
        'trend': 'improving' if progress > 2 else 'declining' if progress < -2 else 'stable',
        'scores': sorted_scores
    }

# ========== 考试管理 ==========

def create_exam(exam_data: Dict[str, Any]) -> Dict[str, Any]:
    """创建考试"""
    exam_id = str(uuid.uuid4())
    
    exam = {
        'id': exam_id,
        'name': exam_data.get('name', '未命名考试'),
        'class_id': exam_data.get('class_id'),
        'grade': exam_data.get('grade'),
        'total_score': float(exam_data.get('total_score', 100)),
        'created_at': datetime.now().isoformat(),
        'status': 'draft',  # draft:草稿, ready:就绪, scanning:扫描中, reviewing:待审核, completed:已完成
        'questions': [],  # 题目信息
        'scores': [],  # 扫描后的成绩
        'statistics': {}
    }
    
    metadata = load_exams_metadata()
    metadata['exams'].append(exam)
    save_exams_metadata(metadata)
    
    return {'success': True, 'exam_id': exam_id, 'exam': exam}

def add_exam_score(exam_id: str, score_data: Dict[str, Any]) -> Dict[str, Any]:
    """添加考试成绩到考试"""
    metadata = load_exams_metadata()
    
    for exam in metadata['exams']:
        if exam['id'] == exam_id:
            # 检查是否已有该学生的成绩
            student_number = score_data.get('student_number')
            for existing in exam['scores']:
                if existing.get('student_number') == student_number:
                    # 更新已有成绩
                    existing.update({
                        'student_name': score_data.get('student_name', existing.get('student_name')),
                        'total_score': float(score_data.get('total_score', 0)),
                        'max_score': float(score_data.get('max_score', exam.get('total_score', 100))),
                        'accuracy': score_data.get('accuracy', 0),
                        'updated_at': datetime.now().isoformat()
                    })
                    save_exams_metadata(metadata)
                    return {'success': True, 'message': '成绩已更新'}
            
            # 添加新成绩
            new_score = {
                'student_number': student_number,
                'student_name': score_data.get('student_name', '未知'),
                'total_score': float(score_data.get('total_score', 0)),
                'max_score': float(score_data.get('max_score', exam.get('total_score', 100))),
                'accuracy': score_data.get('accuracy', 0),
                'student_id': score_data.get('student_id'),
                'confirmed': False,
                'adjusted': False,
                'created_at': datetime.now().isoformat()
            }
            exam['scores'].append(new_score)
            save_exams_metadata(metadata)
            return {'success': True, 'message': '成绩已添加'}
    
    return {'success': False, 'error': '考试不存在'}

def add_question_to_exam(exam_id: str, question_data: Dict[str, Any]) -> bool:
    """添加题目到考试"""
    metadata = load_exams_metadata()
    for exam in metadata['exams']:
        if exam['id'] == exam_id:
            question = {
                'number': len(exam['questions']) + 1,
                'type': question_data.get('type', 'choice'),  # choice, true_false, fill_blank, subjective
                'content': question_data.get('content', ''),
                'correct_answer': question_data.get('correct_answer'),
                'score': float(question_data.get('score', 5)),
                'knowledge_points': question_data.get('knowledge_points', []),
                'difficulty': question_data.get('difficulty', 'medium')
            }
            exam['questions'].append(question)
            save_exams_metadata(metadata)
            return True
    return False

def set_exam_ready(exam_id: str) -> bool:
    """设置考试就绪状态"""
    metadata = load_exams_metadata()
    for exam in metadata['exams']:
        if exam['id'] == exam_id:
            if len(exam['questions']) == 0:
                return False
            exam['status'] = 'ready'
            save_exams_metadata(metadata)
            return True
    return False

def get_exam_by_id(exam_id: str) -> Optional[Dict[str, Any]]:
    """获取考试详情"""
    metadata = load_exams_metadata()
    for exam in metadata['exams']:
        if exam['id'] == exam_id:
            return exam
    return None

def get_exams_by_class(class_id: str) -> List[Dict[str, Any]]:
    """获取班级的所有考试"""
    metadata = load_exams_metadata()
    return [e for e in metadata['exams'] if e.get('class_id') == class_id]

def delete_exam(exam_id: str) -> bool:
    """删除考试"""
    metadata = load_exams_metadata()
    initial_length = len(metadata['exams'])
    metadata['exams'] = [e for e in metadata['exams'] if e['id'] != exam_id]
    
    if len(metadata['exams']) < initial_length:
        save_exams_metadata(metadata)
        return True
    return False

# ========== 答题卡识别（模拟实现）============

def detect_student_number(image_path: str) -> str:
    """识别学生考号"""
    # 实际应该调用OCR/AI来识别
    # 这里返回模拟数据
    return ''

def detect_answers(image_path: str, questions: List[Dict]) -> Dict[int, Any]:
    """识别学生答题结果"""
    # 实际应该调用AI来识别每道题的答案
    # 这里返回模拟数据
    answers = {}
    for q in questions:
        if q['type'] in ['choice', 'true_false']:
            answers[q['number']] = q.get('correct_answer')
    return answers

def scan_answer_sheet(exam_id: str, image_path: str) -> Dict[str, Any]:
    """扫描单张答题卡"""
    exam = get_exam_by_id(exam_id)
    if not exam:
        return {'success': False, 'error': '考试不存在'}
    
    if exam['status'] not in ['ready', 'scanning']:
        return {'success': False, 'error': '考试状态不允许扫描'}
    
    try:
        student_number = detect_student_number(image_path)
        answers = detect_answers(image_path, exam['questions'])
        
        results = []
        correct_count = 0
        total_score = 0
        knowledge_stats = {}
        
        for q in exam['questions']:
            q_num = q['number']
            student_answer = answers.get(q_num)
            is_correct = None
            score = 0
            
            if q['type'] in ['choice', 'true_false', 'fill_blank']:
                if student_answer is not None:
                    is_correct = (student_answer == q['correct_answer'])
                    if is_correct:
                        correct_count += 1
                        score = q['score']
                        total_score += q['score']
                    else:
                        score = 0
                else:
                    is_correct = False
                    score = 0
            
            for kp in q.get('knowledge_points', []):
                if kp not in knowledge_stats:
                    knowledge_stats[kp] = {'total': 0, 'correct': 0}
                knowledge_stats[kp]['total'] += 1
                if is_correct:
                    knowledge_stats[kp]['correct'] += 1
            
            results.append({
                'number': q_num,
                'type': q['type'],
                'student_answer': student_answer,
                'correct_answer': q['correct_answer'],
                'is_correct': is_correct,
                'score': score,
                'max_score': q['score'],
                'knowledge_points': q.get('knowledge_points', [])
            })
        
        return {
            'success': True,
            'student_number': student_number,
            'results': results,
            'total_score': total_score,
            'max_score': exam['total_score'],
            'correct_count': correct_count,
            'total_questions': len(exam['questions']),
            'accuracy': (correct_count / len(exam['questions']) * 100) if exam['questions'] else 0,
            'knowledge_stats': knowledge_stats
        }
    except Exception as e:
        logger.error(f'答题卡扫描失败: {e}')
        return {'success': False, 'error': str(e)}

def batch_scan_answer_sheets(exam_id: str, image_files: List[Tuple[str, bytes]]) -> Dict[str, Any]:
    """批量扫描答题卡"""
    exam = get_exam_by_id(exam_id)
    if not exam:
        return {'success': False, 'error': '考试不存在'}
    
    metadata = load_exams_metadata()
    for e in metadata['exams']:
        if e['id'] == exam_id:
            e['status'] = 'scanning'
            save_exams_metadata(metadata)
            break
    
    cls = get_class_by_id(exam.get('class_id'))
    student_map = {}
    if cls:
        for student in cls.get('students', []):
            student_map[student.get('student_number', '')] = student
    
    scanned_results = []
    matched_count = 0
    unmatched_count = 0
    
    for idx, (filename, image_data) in enumerate(image_files):
        ext = filename.split('.')[-1] if '.' in filename else 'jpg'
        image_path = os.path.join(TEST_LIBRARY_DIR, 'scans', f'{exam_id}_{idx}.{ext}')
        os.makedirs(os.path.dirname(image_path), exist_ok=True)
        with open(image_path, 'wb') as f:
            f.write(image_data)
        
        scan_result = scan_answer_sheet(exam_id, image_path)
        if scan_result.get('success'):
            student_info = student_map.get(scan_result.get('student_number', ''))
            if student_info:
                scan_result['student_id'] = student_info['id']
                scan_result['student_name'] = student_info['name']
                matched_count += 1
            else:
                unmatched_count += 1
                scan_result['student_id'] = None
                scan_result['student_name'] = '未匹配'
            
            scan_result['image_path'] = image_path
            scanned_results.append(scan_result)
    
    for e in metadata['exams']:
        if e['id'] == exam_id:
            e['scores'] = scanned_results
            e['status'] = 'reviewing'
            save_exams_metadata(metadata)
            break
    
    return {
        'success': True,
        'total_scanned': len(scanned_results),
        'matched': matched_count,
        'unmatched': unmatched_count,
        'results': scanned_results
    }

def confirm_exam_scores(exam_id: str) -> Dict[str, Any]:
    """确认考试成绩"""
    exam = get_exam_by_id(exam_id)
    if not exam:
        return {'success': False, 'error': '考试不存在'}
    
    if exam['status'] != 'reviewing':
        return {'success': False, 'error': '考试状态不允许确认'}
    
    metadata = load_exams_metadata()
    for e in metadata['exams']:
        if e['id'] == exam_id:
            scores = [s['total_score'] for s in e['scores'] if s.get('student_id')]
            if scores:
                e['statistics'] = {
                    'total_students': len(scores),
                    'average_score': sum(scores) / len(scores),
                    'highest_score': max(scores),
                    'lowest_score': min(scores),
                    'pass_count': sum(1 for s in scores if s / e['total_score'] >= 0.6),
                    'pass_rate': sum(1 for s in scores if s / e['total_score'] >= 0.6) / len(scores) * 100
                }
            
            if e.get('class_id'):
                for score in e['scores']:
                    if score.get('student_id'):
                        add_student_score(
                            e['class_id'],
                            score['student_id'],
                            {
                                'paper_id': exam_id,
                                'paper_name': e['name'],
                                'score': score['total_score'],
                                'total_score': e['total_score'],
                                'notes': f"答题卡扫描 - {score.get('student_name', '未知')}"
                            }
                        )
            
            e['status'] = 'completed'
            save_exams_metadata(metadata)
            return {'success': True, 'statistics': e['statistics']}
    
    return {'success': False, 'error': '考试不存在'}

def adjust_score(exam_id: str, student_number: str, score: float) -> Dict[str, Any]:
    """调整单条成绩"""
    metadata = load_exams_metadata()
    for exam in metadata['exams']:
        if exam['id'] == exam_id:
            for s in exam['scores']:
                if s.get('student_number') == student_number:
                    s['total_score'] = score
                    s['adjusted'] = True
                    save_exams_metadata(metadata)
                    return {'success': True}
            return {'success': False, 'error': '未找到该学生成绩'}
    return {'success': False, 'error': '考试不存在'}

def batch_confirm_scores(exam_id: str, student_numbers: List[str]) -> Dict[str, Any]:
    """批量确认成绩"""
    metadata = load_exams_metadata()
    confirmed_count = 0
    
    for exam in metadata['exams']:
        if exam['id'] == exam_id:
            for s in exam['scores']:
                if s.get('student_number') in student_numbers:
                    s['confirmed'] = True
                    confirmed_count += 1
            
            if confirmed_count > 0:
                save_exams_metadata(metadata)
                return {
                    'success': True,
                    'confirmed': confirmed_count,
                    'message': f'成功确认 {confirmed_count} 条成绩'
                }
            return {'success': False, 'error': '未找到匹配的学生成绩'}
    return {'success': False, 'error': '考试不存在'}

def batch_adjust_scores(exam_id: str, adjustments: List[Dict[str, Any]]) -> Dict[str, Any]:
    """批量调整成绩"""
    metadata = load_exams_metadata()
    adjusted_count = 0
    failed = []
    
    for exam in metadata['exams']:
        if exam['id'] == exam_id:
            for adjustment in adjustments:
                student_number = adjustment.get('student_number')
                score = adjustment.get('score')
                
                if not student_number or score is None:
                    failed.append({'student_number': student_number, 'error': '参数不完整'})
                    continue
                
                found = False
                for s in exam['scores']:
                    if s.get('student_number') == student_number:
                        s['total_score'] = float(score)
                        s['adjusted'] = True
                        adjusted_count += 1
                        found = True
                        break
                
                if not found:
                    failed.append({'student_number': student_number, 'error': '未找到该学生'})
            
            if adjusted_count > 0:
                save_exams_metadata(metadata)
            
            return {
                'success': True,
                'adjusted': adjusted_count,
                'failed': failed,
                'message': f'成功调整 {adjusted_count} 条成绩' + (f', {len(failed)} 条失败' if failed else '')
            }
    return {'success': False, 'error': '考试不存在'}

def get_exam_analysis(exam_id: str) -> Dict[str, Any]:
    """获取考试详细分析"""
    exam = get_exam_by_id(exam_id)
    if not exam:
        return {'success': False, 'error': '考试不存在'}
    
    question_stats = []
    for q in exam['questions']:
        correct_count = 0
        for score in exam.get('scores', []):
            for r in score.get('results', []):
                if r['number'] == q['number'] and r.get('is_correct'):
                    correct_count += 1
        
        total = len(exam.get('scores', []))
        question_stats.append({
            'number': q['number'],
            'type': q['type'],
            'knowledge_points': q.get('knowledge_points', []),
            'correct_count': correct_count,
            'error_count': total - correct_count,
            'error_rate': ((total - correct_count) / total * 100) if total > 0 else 0
        })
    
    knowledge_stats = {}
    for q in exam['questions']:
        for kp in q.get('knowledge_points', []):
            if kp not in knowledge_stats:
                knowledge_stats[kp] = {'total': 0, 'correct': 0}
            knowledge_stats[kp]['total'] += len(exam.get('scores', []))
    
    for score in exam.get('scores', []):
        for r in score.get('results', []):
            for kp in r.get('knowledge_points', []):
                if kp in knowledge_stats and r.get('is_correct'):
                    knowledge_stats[kp]['correct'] += 1
    
    for kp, stats in knowledge_stats.items():
        stats['mastery_rate'] = (stats['correct'] / stats['total'] * 100) if stats['total'] > 0 else 0
    
    return {
        'success': True,
        'exam': exam,
        'question_stats': question_stats,
        'knowledge_stats': knowledge_stats,
        'statistics': exam.get('statistics', {})
    }

# ========== 批量上传分析（兼容旧接口）==========

def get_all_wrong_questions(paper_ids: List[str] = None, grade: str = None) -> List[Dict[str, Any]]:
    """获取错题"""
    papers = get_all_papers()
    if paper_ids:
        papers = [p for p in papers if p['id'] in paper_ids]
    if grade:
        papers = [p for p in papers if p.get('grade') == grade]
    
    wrong_questions = []
    for paper in papers:
        analysis = load_paper_analysis(paper.get('id'))
        if not analysis:
            continue
        for q in analysis.get('questions', []):
            if not q.get('is_correct', True):
                q_copy = q.copy()
                q_copy['paper_id'] = paper.get('id')
                q_copy['paper_name'] = paper.get('name', '未命名试卷')
                wrong_questions.append(q_copy)
    
    return wrong_questions

def get_wrong_questions(paper_ids: List[str] = None, grade: str = None) -> List[Dict[str, Any]]:
    """获取错题（兼容现有API）"""
    return get_all_wrong_questions(paper_ids, grade)

def generate_prompt_optimization_suggestions(paper_ids: List[str] = None, grade: str = None) -> Dict[str, Any]:
    """生成提示词优化建议"""
    papers = get_all_papers()
    if paper_ids:
        papers = [p for p in papers if p['id'] in paper_ids]
    if grade:
        papers = [p for p in papers if p.get('grade') == grade]
    
    all_questions = []
    for paper in papers:
        analysis = load_paper_analysis(paper.get('id'))
        if analysis:
            all_questions.extend(analysis.get('questions', []))
    
    if not all_questions:
        return {'success': True, 'suggestions': ['暂无数据，请先上传试卷'], 'stats': {}}
    
    total_questions = len(all_questions)
    correct_count = sum(1 for q in all_questions if q.get('is_correct', False))
    accuracy = correct_count / total_questions * 100 if total_questions > 0 else 0
    
    suggestions = []
    if accuracy < 60:
        suggestions.append({'type': '整体建议', 'priority': 'high', 'suggestion': '整体正确率较低，建议增加基础知识点的讲解'})
    elif accuracy < 80:
        suggestions.append({'type': '整体建议', 'priority': 'medium', 'suggestion': '整体情况尚可，建议针对薄弱知识点进行专项练习'})
    else:
        suggestions.append({'type': '整体建议', 'priority': 'low', 'suggestion': '整体掌握良好，可以适当增加拓展内容'})
    
    return {'success': True, 'suggestions': suggestions, 'stats': {'total_papers': len(papers), 'total_questions': total_questions, 'accuracy': round(accuracy, 1)}}

def batch_analyze_papers(image_files: List[Tuple[str, bytes]], 
                         grade: str, 
                         paper_name: str,
                         concurrency: int = 4,
                         class_id: str = None,
                         auto_detect_names: bool = True,
                         work_mode: str = 'auto') -> Dict[str, Any]:
    """批量分析试卷（兼容旧接口）"""
    paper_id = str(uuid.uuid4())
    
    metadata = load_library_metadata()
    paper_metadata = {
        'id': paper_id,
        'name': paper_name,
        'grade': grade,
        'upload_time': datetime.now().isoformat(),
        'image_count': len(image_files),
        'status': 'completed',
        'questions': [],
        'knowledge_points': [],
        'work_mode': work_mode
    }
    metadata['papers'].append(paper_metadata)
    save_library_metadata(metadata)
    
    for idx, (filename, image_data) in enumerate(image_files):
        ext = filename.split('.')[-1] if '.' in filename else 'jpg'
        image_path = os.path.join(TEST_LIBRARY_DIR, 'images', f'{paper_id}_{idx}.{ext}')
        with open(image_path, 'wb') as f:
            f.write(image_data)
    
    return {
        'success': True,
        'paper_id': paper_id,
        'total_questions': 0,
        'detected_students': 0
    }

def delete_paper(paper_id: str) -> bool:
    """删除试卷"""
    metadata = load_library_metadata()
    initial_length = len(metadata['papers'])
    metadata['papers'] = [p for p in metadata['papers'] if p['id'] != paper_id]
    
    if len(metadata['papers']) < initial_length:
        save_library_metadata(metadata)
        return True
    return False

# ========== 知识图谱 ==========

def build_knowledge_point_graph() -> Dict[str, Any]:
    """构建知识点关联图谱"""
    papers = get_all_papers()
    knowledge_relations = {}
    knowledge_mastery = {}
    
    for paper in papers:
        analysis = load_paper_analysis(paper.get('id'))
        if not analysis:
            continue
        
        for question in analysis.get('questions', []):
            for kp in question.get('knowledge_points', []):
                if kp not in knowledge_relations:
                    knowledge_relations[kp] = {'total_questions': 0, 'correct_count': 0, 'students': []}
                
                knowledge_relations[kp]['total_questions'] += 1
                if question.get('is_correct', False):
                    knowledge_relations[kp]['correct_count'] += 1
    
    for kp, data in knowledge_relations.items():
        if data['total_questions'] > 0:
            knowledge_mastery[kp] = {
                'mastery_rate': data['correct_count'] / data['total_questions'] * 100,
                'total_questions': data['total_questions'],
                'correct_count': data['correct_count']
            }
    
    return {'success': True, 'knowledge_points': knowledge_mastery, 'relations': knowledge_relations}

# ========== 班级学情分析 ==========

def analyze_class_performance(class_id: str) -> Dict[str, Any]:
    """分析班级整体学情"""
    cls = get_class_by_id(class_id)
    if not cls:
        return {'success': False, 'error': '班级不存在'}
    
    scores = get_class_scores(class_id)
    
    if not scores:
        return {'success': True, 'class_id': class_id, 'data': None}
    
    score_values = [s['score'] for s in scores]
    percentages = [s['score'] / s['total_score'] * 100 if s['total_score'] > 0 else 0 for s in scores]
    
    return {
        'success': True,
        'class_id': class_id,
        'class_name': cls['name'],
        'total_students': len(cls.get('students', [])),
        'participated': len(scores),
        'statistics': {
            'average_score': round(sum(score_values) / len(score_values), 1) if score_values else 0,
            'average_percentage': round(sum(percentages) / len(percentages), 1) if percentages else 0,
            'highest_score': max(score_values) if score_values else 0,
            'lowest_score': min(score_values) if score_values else 0,
            'pass_rate': round(sum(1 for p in percentages if p >= 60) / len(percentages) * 100, 1) if percentages else 0
        }
    }

def generate_class_report(class_id: str) -> Dict[str, Any]:
    """生成班级报告"""
    performance = analyze_class_performance(class_id)
    if not performance.get('success'):
        return performance
    
    statistics = performance.get('statistics', {})
    
    suggestions = []
    if statistics.get('average_percentage', 0) < 60:
        suggestions.append('班级整体成绩偏低，建议加强基础教学')
    if statistics.get('pass_rate', 0) < 70:
        suggestions.append('及格率有待提高，关注中下游学生')
    
    report = {
        'generated_at': datetime.now().isoformat(),
        'class_id': class_id,
        'class_name': performance.get('class_name'),
        'statistics': statistics,
        'suggestions': suggestions,
        'participated': performance.get('participated', 0)
    }
    
    # 保存报告
    reports_dir = os.path.join(REPORTS_DIR, class_id)
    os.makedirs(reports_dir, exist_ok=True)
    report_path = os.path.join(reports_dir, f'report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json')
    save_json(report_path, report)
    
    return {'success': True, 'report': report}

# ========== 成绩导出功能 ==========

def export_exam_scores(exam_id: str, export_format: str = 'csv') -> Optional[str]:
    """
    导出考试成绩为CSV或Excel格式
    
    Args:
        exam_id: 考试ID
        export_format: 导出格式 ('csv' 或 'excel')
    
    Returns:
        导出文件的临时路径，失败返回None
    """
    exam = get_exam_by_id(exam_id)
    if not exam:
        logger.error(f'考试不存在: {exam_id}')
        return None
    
    scores = exam.get('scores', [])
    if not scores:
        logger.warning(f'考试没有成绩数据: {exam_id}')
        return None
    
    # 创建导出目录
    export_dir = os.path.join(DATA_DIR, 'exports')
    os.makedirs(export_dir, exist_ok=True)
    
    # 生成文件名
    exam_name = exam.get('name', '未知考试').replace(' ', '_').replace('/', '_')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'{exam_name}_{timestamp}.{export_format}'
    filepath = os.path.join(export_dir, filename)
    
    if export_format == 'csv':
        return _export_to_csv(exam, scores, filepath)
    elif export_format == 'excel':
        return _export_to_excel(exam, scores, filepath)
    else:
        logger.error(f'不支持的导出格式: {export_format}')
        return None

def _export_to_csv(exam: Dict, scores: List[Dict], filepath: str) -> Optional[str]:
    """导出为CSV格式"""
    try:
        import csv
        
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            
            # 写入标题
            writer.writerow([
                '学号', '姓名', '总分', '满分', '正确率(%)', 
                '匹配状态', '确认状态', '调整标记',
                '考试名称', '考试日期', '班级'
            ])
            
            # 写入数据
            for score in scores:
                writer.writerow([
                    score.get('student_number', ''),
                    score.get('student_name', '未匹配'),
                    score.get('total_score', 0),
                    score.get('max_score', 100),
                    round(score.get('accuracy', 0), 1),
                    '已匹配' if score.get('student_id') else '未匹配',
                    '已确认' if score.get('confirmed') else '未确认',
                    '已调整' if score.get('adjusted') else '未调整',
                    exam.get('name', ''),
                    exam.get('created_at', ''),
                    exam.get('class_name', '')
                ])
        
        logger.info(f'CSV导出成功: {filepath}')
        return filepath
    except Exception as e:
        logger.error(f'CSV导出失败: {e}')
        return None

def _export_to_excel(exam: Dict, scores: List[Dict], filepath: str) -> Optional[str]:
    """导出为Excel格式"""
    try:
        # 尝试导入openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning('openpyxl未安装，尝试安装...')
            import subprocess
            subprocess.check_call(['pip', 'install', 'openpyxl', '-q'])
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = '成绩单'
        
        # 定义样式
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 标题行
        headers = ['学号', '姓名', '总分', '满分', '正确率(%)', 
                   '匹配状态', '确认状态', '调整标记', '考试名称', '考试日期', '班级']
        ws.append(headers)
        
        # 设置标题样式
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 数据行
        for score in scores:
            ws.append([
                score.get('student_number', ''),
                score.get('student_name', '未匹配'),
                score.get('total_score', 0),
                score.get('max_score', 100),
                round(score.get('accuracy', 0), 1),
                '已匹配' if score.get('student_id') else '未匹配',
                '已确认' if score.get('confirmed') else '未确认',
                '已调整' if score.get('adjusted') else '未调整',
                exam.get('name', ''),
                exam.get('created_at', ''),
                exam.get('class_name', '')
            ])
        
        # 自动调整列宽
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            for row in range(2, ws.max_row + 1):
                try:
                    if ws.cell(row=row, column=col).value:
                        max_length = max(max_length, len(str(ws.cell(row=row, column=col).value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # 添加统计信息sheet
        ws_stats = wb.create_sheet(title='统计信息')
        
        # 计算统计数据
        total_students = len(scores)
        matched_students = sum(1 for s in scores if s.get('student_id'))
        confirmed_students = sum(1 for s in scores if s.get('confirmed'))
        adjusted_students = sum(1 for s in scores if s.get('adjusted'))
        
        total_scores = [s.get('total_score', 0) for s in scores]
        avg_score = sum(total_scores) / len(total_scores) if total_scores else 0
        max_score = max(total_scores) if total_scores else 0
        min_score = min(total_scores) if total_scores else 0
        
        accuracies = [s.get('accuracy', 0) for s in scores]
        avg_accuracy = sum(accuracies) / len(accuracies) if accuracies else 0
        
        # 写入统计信息
        stats_data = [
            ['项目', '数值'],
            ['考试名称', exam.get('name', '')],
            ['考试日期', exam.get('created_at', '')],
            ['班级', exam.get('class_name', '')],
            ['总人数', total_students],
            ['已匹配人数', matched_students],
            ['已确认人数', confirmed_students],
            ['已调整人数', adjusted_students],
            ['平均分', round(avg_score, 2)],
            ['最高分', max_score],
            ['最低分', min_score],
            ['平均正确率(%)', round(avg_accuracy, 2)],
            ['', ''],
            ['导出时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        for row_data in stats_data:
            ws_stats.append(row_data)
        
        # 设置统计信息sheet样式
        ws_stats.cell(row=1, column=1).font = Font(bold=True)
        ws_stats.cell(row=1, column=2).font = Font(bold=True)
        ws_stats.column_dimensions['A'].width = 20
        ws_stats.column_dimensions['B'].width = 30
        
        wb.save(filepath)
        logger.info(f'Excel导出成功: {filepath}')
        return filepath
    except Exception as e:
        logger.error(f'Excel导出失败: {e}')
        return None

def export_class_scores(class_id: str, export_format: str = 'csv') -> Optional[str]:
    """
    导出班级所有考试成绩
    
    Args:
        class_id: 班级ID
        export_format: 导出格式
    
    Returns:
        导出文件的临时路径，失败返回None
    """
    cls = get_class_by_id(class_id)
    if not cls:
        logger.error(f'班级不存在: {class_id}')
        return None
    
    exams = get_exams_by_class(class_id)
    
    # 创建导出目录
    export_dir = os.path.join(DATA_DIR, 'exports')
    os.makedirs(export_dir, exist_ok=True)
    
    # 生成文件名
    class_name = cls.get('name', '未知班级').replace(' ', '_').replace('/', '_')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'{class_name}_成绩汇总_{timestamp}.{export_format}'
    filepath = os.path.join(export_dir, filename)
    
    if export_format == 'csv':
        return _export_class_to_csv(cls, exams, filepath)
    elif export_format == 'excel':
        return _export_class_to_excel(cls, exams, filepath)
    else:
        return None

def _export_class_to_csv(cls: Dict, exams: List[Dict], filepath: str) -> Optional[str]:
    """导出班级成绩为CSV"""
    try:
        import csv
        
        # 收集所有考试名称
        exam_names = [e.get('name', '未知考试') for e in exams if e.get('status') == 'completed']
        
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            
            # 标题行
            headers = ['学号', '姓名'] + exam_names + ['平均分', '总分']
            writer.writerow(headers)
            
            # 按学生分组
            students_data = {}
            for exam in exams:
                if exam.get('status') != 'completed':
                    continue
                for score in exam.get('scores', []):
                    student_number = score.get('student_number', '')
                    if student_number:
                        if student_number not in students_data:
                            students_data[student_number] = {
                                'name': score.get('student_name', ''),
                                'scores': {}
                            }
                        students_data[student_number]['scores'][exam.get('name')] = score.get('total_score', 0)
            
            # 写入学生数据
            for student_number, data in students_data.items():
                row = [student_number, data['name']]
                total = 0
                count = 0
                for exam_name in exam_names:
                    score = data['scores'].get(exam_name, '-')
                    row.append(score)
                    if isinstance(score, (int, float)):
                        total += score
                        count += 1
                row.append(round(total / count, 1) if count > 0 else '-')
                row.append(total if count > 0 else '-')
                writer.writerow(row)
        
        return filepath
    except Exception as e:
        logger.error(f'班级成绩CSV导出失败: {e}')
        return None

def _export_class_to_excel(cls: Dict, exams: List[Dict], filepath: str) -> Optional[str]:
    """导出班级成绩为Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = '成绩汇总'
        
        # 收集所有完成的考试
        completed_exams = [e for e in exams if e.get('status') == 'completed']
        exam_names = [e.get('name', '未知考试') for e in completed_exams]
        
        # 标题样式
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        # 标题行
        headers = ['学号', '姓名'] + exam_names + ['平均分', '总分']
        ws.append(headers)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # 按学生分组
        students_data = {}
        for exam in completed_exams:
            for score in exam.get('scores', []):
                student_number = score.get('student_number', '')
                if student_number:
                    if student_number not in students_data:
                        students_data[student_number] = {
                            'name': score.get('student_name', ''),
                            'scores': {}
                        }
                    students_data[student_number]['scores'][exam.get('name')] = score.get('total_score', 0)
        
        # 写入学生数据
        for student_number, data in students_data.items():
            row = [student_number, data['name']]
            total = 0
            count = 0
            for exam_name in exam_names:
                score = data['scores'].get(exam_name, '-')
                row.append(score)
                if isinstance(score, (int, float)):
                    total += score
                    count += 1
            row.append(round(total / count, 1) if count > 0 else '-')
            row.append(total if count > 0 else '-')
            ws.append(row)
        
        # 自动调整列宽
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            for row in range(2, ws.max_row + 1):
                try:
                    if ws.cell(row=row, column=col).value:
                        max_length = max(max_length, len(str(ws.cell(row=row, column=col).value)))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 30)
        
        wb.save(filepath)
        return filepath
    except Exception as e:
        logger.error(f'班级成绩Excel导出失败: {e}')
        return None

# ========== 学生端成绩查询 ==========

def get_student_scores_by_number(student_number: str) -> Optional[Dict[str, Any]]:
    """
    根据学号获取学生所有成绩
    
    Args:
        student_number: 学生学号
    
    Returns:
        学生信息和成绩列表
    """
    metadata = load_exams_metadata()
    student_scores = []
    student_info = None
    
    # 遍历所有考试，收集该学生的成绩
    for exam in metadata['exams']:
        for score in exam.get('scores', []):
            if score.get('student_number') == student_number:
                # 获取学生基本信息（从第一个匹配的考试中获取）
                if not student_info:
                    student_info = {
                        'student_number': student_number,
                        'name': score.get('student_name', '未知'),
                        'class_id': exam.get('class_id'),
                        'class_name': exam.get('class_name', '')
                    }
                
                student_scores.append({
                    'exam_id': exam['id'],
                    'exam_name': exam.get('name', '未知考试'),
                    'score': score.get('total_score', 0),
                    'max_score': score.get('max_score', 100),
                    'accuracy': score.get('accuracy', 0),
                    'date': exam.get('created_at', ''),
                    'status': exam.get('status', ''),
                    'confirmed': score.get('confirmed', False),
                    'adjusted': score.get('adjusted', False)
                })
    
    if not student_info:
        return None
    
    # 计算概览数据
    overview = calculate_student_overview(student_scores)
    
    # 获取知识点统计
    knowledge_stats = calculate_student_knowledge_stats(student_number, metadata)
    
    return {
        'success': True,
        'student': student_info,
        'exams': student_scores,
        'overview': overview,
        'knowledge_stats': knowledge_stats
    }

def calculate_student_overview(scores: List[Dict]) -> Dict[str, Any]:
    """计算学生成绩概览"""
    if not scores:
        return {
            'total_exams': 0,
            'avg_score': 0,
            'best_score': 0,
            'avg_accuracy': 0,
            'chart_data': []
        }
    
    # 按日期排序
    sorted_scores = sorted(scores, key=lambda x: x.get('date', ''))
    
    total_exams = len(scores)
    avg_score = sum(s['score'] for s in scores) / total_exams
    best_score = max(s['score'] for s in scores)
    avg_accuracy = sum(s['accuracy'] for s in scores) / total_exams
    
    # 图表数据
    chart_data = []
    for i, score in enumerate(sorted_scores[-10:]):  # 只取最近10次
        chart_data.append({
            'label': f'考试{i+1}',
            'score': score['score'],
            'max_score': score['max_score']
        })
    
    return {
        'total_exams': total_exams,
        'avg_score': round(avg_score, 1),
        'best_score': round(best_score, 1),
        'avg_accuracy': round(avg_accuracy, 3),
        'chart_data': chart_data
    }

def calculate_student_knowledge_stats(student_number: str, metadata: Dict) -> Dict[str, Any]:
    """计算学生知识点掌握情况"""
    knowledge_stats = defaultdict(lambda: {'correct': 0, 'total': 0})
    
    for exam in metadata['exams']:
        for score in exam.get('scores', []):
            if score.get('student_number') == student_number:
                # 获取该考试的题目知识点信息
                questions = exam.get('questions', [])
                results = score.get('results', [])
                
                for result in results:
                    q_number = result.get('number')
                    # 找到对应的题目知识点
                    for q in questions:
                        if q.get('number') == q_number:
                            kps = q.get('knowledge_points', [])
                            if not kps:
                                kps = ['基础题']
                            
                            for kp in kps:
                                knowledge_stats[kp]['total'] += 1
                                if result.get('is_correct'):
                                    knowledge_stats[kp]['correct'] += 1
                            break
    
    # 计算掌握度
    result = {}
    for kp, stats in knowledge_stats.items():
        mastery_rate = stats['correct'] / stats['total'] if stats['total'] > 0 else 0
        result[kp] = {
            'correct': stats['correct'],
            'total': stats['total'],
            'mastery_rate': mastery_rate
        }
    
    return result

def get_student_exam_detail(student_number: str, exam_id: str) -> Optional[Dict[str, Any]]:
    """获取学生某次考试的详细信息"""
    exam = get_exam_by_id(exam_id)
    if not exam:
        return None
    
    # 找到该学生的成绩
    student_score = None
    for score in exam.get('scores', []):
        if score.get('student_number') == student_number:
            student_score = score
            break
    
    if not student_score:
        return None
    
    # 获取班级排名
    all_scores = sorted(
        exam.get('scores', []),
        key=lambda x: x.get('total_score', 0),
        reverse=True
    )
    rank = None
    for i, s in enumerate(all_scores):
        if s.get('student_number') == student_number:
            rank = i + 1
            break
    
    # 获取题目详情
    questions = exam.get('questions', [])
    results = student_score.get('results', [])
    
    question_details = []
    for result in results:
        q_number = result.get('number')
        q_info = next((q for q in questions if q.get('number') == q_number), {})
        
        question_details.append({
            'number': q_number,
            'type': q_info.get('type', '未知'),
            'is_correct': result.get('is_correct', False),
            'knowledge_points': q_info.get('knowledge_points', [])
        })
    
    return {
        'success': True,
        'exam_name': exam.get('name', ''),
        'score': student_score.get('total_score', 0),
        'max_score': student_score.get('max_score', 100),
        'accuracy': student_score.get('accuracy', 0),
        'rank': rank,
        'total_students': len(all_scores),
        'confirmed': student_score.get('confirmed', False),
        'adjusted': student_score.get('adjusted', False),
        'questions': question_details
    }

# ========== 错题本功能 ==========

def get_student_wrong_questions(student_number: str) -> Dict[str, Any]:
    """
    获取学生的错题本
    
    Args:
        student_number: 学号
    
    Returns:
        错题列表，按知识点分类
    """
    metadata = load_exams_metadata()
    wrong_questions = []
    
    # 遍历所有考试，收集该学生的错题
    for exam in metadata['exams']:
        for score in exam.get('scores', []):
            if score.get('student_number') == student_number:
                exam_name = exam.get('name', '未知考试')
                exam_date = exam.get('created_at', '')
                questions = exam.get('questions', [])
                results = score.get('results', [])
                
                # 收集错题
                for result in results:
                    if not result.get('is_correct', True):  # 错题
                        q_number = result.get('number')
                        
                        # 找到对应的题目信息
                        q_info = None
                        for q in questions:
                            if q.get('number') == q_number:
                                q_info = q
                                break
                        
                        if q_info:
                            wrong_questions.append({
                                'exam_id': exam['id'],
                                'exam_name': exam_name,
                                'exam_date': exam_date,
                                'question_number': q_number,
                                'question_content': q_info.get('content', ''),
                                'question_type': q_info.get('type', '未知'),
                                'knowledge_points': q_info.get('knowledge_points', ['基础题']),
                                'correct_answer': q_info.get('correct_answer', ''),
                                'student_answer': result.get('student_answer', ''),
                                'score': result.get('score', 0),
                                'max_score': q_info.get('score', 5),
                                'analysis': q_info.get('analysis', ''),
                                'is_mastered': result.get('is_mastered', False),  # 是否已掌握
                                'practice_count': result.get('practice_count', 0),  # 练习次数
                                'last_practice_date': result.get('last_practice_date')
                            })
    
    # 按知识点分类
    knowledge_groups = defaultdict(list)
    for wq in wrong_questions:
        for kp in wq['knowledge_points']:
            knowledge_groups[kp].append(wq)
    
    # 统计知识点掌握情况
    knowledge_mastery = {}
    for kp, questions in knowledge_groups.items():
        total = len(questions)
        mastered = sum(1 for q in questions if q.get('is_mastered'))
        knowledge_mastery[kp] = {
            'total': total,
            'mastered': mastered,
            'remaining': total - mastered,
            'mastery_rate': mastered / total if total > 0 else 0
        }
    
    return {
        'success': True,
        'student_number': student_number,
        'total_wrong': len(wrong_questions),
        'total_mastered': sum(1 for q in wrong_questions if q.get('is_mastered')),
        'wrong_questions': wrong_questions,
        'knowledge_groups': dict(knowledge_groups),
        'knowledge_mastery': knowledge_mastery
    }

def mark_question_mastered(student_number: str, exam_id: str, question_number: int, mastered: bool = True) -> Dict[str, Any]:
    """
    标记题目已掌握
    
    Args:
        student_number: 学号
        exam_id: 考试ID
        question_number: 题目编号
        mastered: 是否已掌握
    
    Returns:
        操作结果
    """
    metadata = load_exams_metadata()
    
    for exam in metadata['exams']:
        if exam['id'] == exam_id:
            for score in exam.get('scores', []):
                if score.get('student_number') == student_number:
                    results = score.get('results', [])
                    for result in results:
                        if result.get('number') == question_number:
                            result['is_mastered'] = mastered
                            result['last_practice_date'] = datetime.now().isoformat()
                            if mastered:
                                result['practice_count'] = result.get('practice_count', 0) + 1
                    
                    save_exams_metadata(metadata)
                    return {'success': True, 'message': '标记成功'}
    
    return {'success': False, 'error': '未找到该题目'}

def get_practice_questions(student_number: str, knowledge_point: str = None, limit: int = 10) -> Dict[str, Any]:
    """
    获取练习题目（从错题中抽取）
    
    Args:
        student_number: 学号
        knowledge_point: 知识点（可选）
        limit: 题目数量
    
    Returns:
        待练习的错题列表
    """
    wrong_data = get_student_wrong_questions(student_number)
    
    if not wrong_data['success']:
        return wrong_data
    
    # 筛选未掌握的错题
    unpracticed = [q for q in wrong_data['wrong_questions'] if not q.get('is_mastered')]
    
    # 如果指定了知识点，过滤
    if knowledge_point:
        unpracticed = [q for q in unpracticed if knowledge_point in q['knowledge_points']]
    
    # 按练习次数和时间排序（优先练习未练习的）
    unpracticed.sort(key=lambda x: (x.get('practice_count', 0), x.get('last_practice_date', '')))
    
    return {
        'success': True,
        'total': len(unpracticed),
        'practice_questions': unpracticed[:limit],
        'knowledge_point': knowledge_point
    }

def get_wrong_question_detail(student_number: str, exam_id: str, question_number: int) -> Optional[Dict[str, Any]]:
    """
    获取错题详情
    
    Args:
        student_number: 学号
        exam_id: 考试ID
        question_number: 题目编号
    
    Returns:
        错题详情
    """
    exam = get_exam_by_id(exam_id)
    if not exam:
        return None
    
    # 找到该学生的成绩
    student_score = None
    for score in exam.get('scores', []):
        if score.get('student_number') == student_number:
            student_score = score
            break
    
    if not student_score:
        return None
    
    # 找到错题
    questions = exam.get('questions', [])
    results = student_score.get('results', [])
    
    q_info = None
    q_result = None
    
    for q in questions:
        if q.get('number') == question_number:
            q_info = q
            break
    
    for result in results:
        if result.get('number') == question_number:
            q_result = result
            break
    
    if not q_info or not q_result:
        return None
    
    return {
        'success': True,
        'exam_id': exam.get('id', ''),
        'exam_name': exam.get('name', ''),
        'exam_date': exam.get('created_at', ''),
        'question': {
            'number': q_info.get('number'),
            'content': q_info.get('content', ''),
            'type': q_info.get('type', '未知'),
            'correct_answer': q_info.get('correct_answer', ''),
            'analysis': q_info.get('analysis', ''),
            'knowledge_points': q_info.get('knowledge_points', ['基础题']),
            'max_score': q_info.get('score', 5)
        },
        'student_answer': {
            'answer': q_result.get('student_answer', ''),
            'is_correct': q_result.get('is_correct', False),
            'score': q_result.get('score', 0),
            'is_mastered': q_result.get('is_mastered', False),
            'practice_count': q_result.get('practice_count', 0)
        }
    }



